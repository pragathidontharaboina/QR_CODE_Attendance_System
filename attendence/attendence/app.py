from flask import Flask, render_template, request, redirect, url_for, session, send_file
import qrcode
import random
import uuid
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Constants
STUDENT_DB = "students.xlsx"
ATTENDANCE_DB = "attendance.xlsx"
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

# ----------- Utility Functions -----------

def get_mac():
    mac = uuid.getnode()
    return ':'.join(("%012X" % mac)[i:i+2] for i in range(0, 12, 2))

def is_student_registered(name, roll):
    if not os.path.exists(STUDENT_DB):
        return False
    wb = load_workbook(STUDENT_DB)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == roll:
            return True
    return False

def is_mac_already_logged(mac):
    if not os.path.exists(ATTENDANCE_DB):
        return False
    wb = load_workbook(ATTENDANCE_DB)
    ws = wb.active
    today = datetime.today().strftime('%Y-%m-%d')
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == today and row[4] == mac:
            return True
    return False

def save_attendance(name, roll, otp, mac):
    if not os.path.exists(ATTENDANCE_DB):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Name", "Roll No", "OTP", "MAC Address"])
    else:
        wb = load_workbook(ATTENDANCE_DB)
        ws = wb.active
    today = datetime.today().strftime('%Y-%m-%d')
    ws.append([today, name, roll, otp, mac])
    wb.save(ATTENDANCE_DB)

# ----------- Routes -----------

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/attendance', methods=['GET', 'POST'])
def attendance():
    if request.method == 'POST':
        name = request.form['name']
        roll = request.form['roll']

        if not is_student_registered(name, roll):
            return render_template('attendance.html', error="Student not registered!")

        otp = str(random.randint(100000, 999999))
        session['otp'] = otp
        session['name'] = name
        session['roll'] = roll

        img_io = BytesIO()
        img = qrcode.make(otp)
        img.save(img_io, 'PNG')
        img_io.seek(0)
        session['qr'] = img_io.read()

        return render_template('attendance.html', otp_mode=True)

    return render_template('attendance.html')

@app.route('/verify_otp', methods=['POST'])
def verify_otp():
    entered_otp = request.form['otp']
    if entered_otp == session.get('otp'):
        mac = get_mac()
        if is_mac_already_logged(mac):
            return render_template('attendance.html', message="MAC already logged today!", otp_mode=False)
        save_attendance(session['name'], session['roll'], session['otp'], mac)
        return render_template('attendance.html', message="Attendance marked successfully!", otp_mode=False)
    else:
        return render_template('attendance.html', error="Incorrect OTP!", otp_mode=True)

@app.route('/qr_image')
def qr_image():
    if 'qr' not in session:
        return "QR not found", 404
    return send_file(BytesIO(session['qr']), mimetype='image/png')

# ----------- Admin Routes -----------

@app.route('/admin', methods=['GET', 'POST'])  # UPDATED route here
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            error = "Invalid credentials. Please try again."
            return render_template('admin_login.html', error=error)
    return render_template('admin_login.html')

@app.route('/admin_dashboard')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    attendance_data = []
    if os.path.exists(ATTENDANCE_DB):
        wb = load_workbook(ATTENDANCE_DB)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            attendance_data.append(row)

    return render_template('admin_dashboard.html', data=attendance_data)

@app.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

# ----------- Run App -----------

if __name__ == '__main__':
    app.run(debug=True)
